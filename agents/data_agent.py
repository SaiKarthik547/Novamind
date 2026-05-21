"""
Data Agent — Full data manipulation, analysis, and transformation engine.
CSV, Excel, JSON, SQL (SQLite + any SQLAlchemy DSN), Parquet, statistical analysis,
data cleaning, pivot tables, chart generation, ETL pipelines, schema inference.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
import sqlite3
import statistics
import tempfile
import uuid
from collections import defaultdict, Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from core.base_agent import BaseAgent

logger = logging.getLogger("DataAgent")
import ast as _ast

# ── Safe formula evaluator (replaces raw eval) ─────────────────────────────────────

_SAFE_AST_NODES = frozenset({
    _ast.Expression, _ast.BinOp, _ast.UnaryOp, _ast.Compare,
    _ast.BoolOp, _ast.Constant, _ast.Name, _ast.Load,
    _ast.Add, _ast.Sub, _ast.Mult, _ast.Div, _ast.Mod, _ast.Pow,
    _ast.FloorDiv, _ast.USub, _ast.UAdd,
    _ast.Eq, _ast.NotEq, _ast.Lt, _ast.LtE, _ast.Gt, _ast.GtE,
    _ast.And, _ast.Or, _ast.Not,
    _ast.IfExp,   # ternary: x if cond else y
})


def _safe_eval_formula(formula: str, row: dict) -> Any:
    """
    Evaluate a simple row-level formula safely.
    Only permits arithmetic, comparisons, and row field references.
    Rejects: imports, function calls, attribute access, arbitrary builtins.
    Returns None on any error.
    """
    try:
        tree = _ast.parse(formula, mode='eval')
        for node in _ast.walk(tree):
            if type(node) not in _SAFE_AST_NODES:
                raise ValueError(f"Unsafe expression node: {type(node).__name__}")
        # Only allow names that match actual row column names
        allowed_names = set(row.keys())
        for node in _ast.walk(tree):
            if isinstance(node, _ast.Name) and node.id not in allowed_names:
                raise ValueError(f"Unknown variable in formula: {node.id!r}")
        return eval(  # noqa: S307  — input already AST-validated above
            compile(tree, '<formula>', 'eval'),
            {"__builtins__": {}},
            row,
        )
    except Exception as _fe:
        logger.warning(f"Formula eval failed: {_fe}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
#  Data Classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class DataProfile:
    rows:           int
    columns:        int
    dtypes:         Dict[str, str]
    null_counts:    Dict[str, int]
    unique_counts:  Dict[str, int]
    numeric_stats:  Dict[str, Dict]   # mean/std/min/max/median
    sample:         List[Dict]         # first 5 rows
    memory_kb:      float


# ─────────────────────────────────────────────────────────────────────────────
#  Data Agent
# ─────────────────────────────────────────────────────────────────────────────

class DataAgent(BaseAgent):
    """
    Full data operations agent.
    Works natively with Python builtins (no pandas required)
    but uses pandas/numpy/openpyxl when available for more power.
    """

    def __init__(self):
        super().__init__()
        self._pd   = self._try_import("pandas")
        self._np   = self._try_import("numpy")
        self._plt  = self._try_import("matplotlib.pyplot")
        self._sns  = self._try_import("seaborn")
        self._xl   = self._try_import("openpyxl")
        self._xl2  = self._try_import("xlrd")
        self._sa   = self._try_import("sqlalchemy")
        self._active_db: Optional[sqlite3.Connection] = None
        self._db_path:   Optional[str]                = None
        self._query_log: List[Dict]                   = []

        self.handlers = {
            # CSV
            "read_csv":           self.read_csv,
            "write_csv":          self.write_csv,
            "append_csv":         self.append_csv,
            "merge_csv":          self.merge_csv,
            "filter_csv":         self.filter_csv,
            "sort_csv":           self.sort_csv,
            "deduplicate_csv":    self.deduplicate_csv,
            "validate_csv":       self.validate_csv,
            "convert_csv_to_json": self.csv_to_json,
            # Excel
            "read_excel":         self.read_excel,
            "write_excel":        self.write_excel,
            "list_sheets":        self.list_excel_sheets,
            "add_sheet":          self.add_excel_sheet,
            "read_sheet":         self.read_excel_sheet,
            "excel_to_csv":       self.excel_to_csv,
            "csv_to_excel":       self.csv_to_excel,
            # JSON
            "read_json":          self.read_json,
            "write_json":         self.write_json,
            "flatten_json":       self.flatten_json,
            "json_to_csv":        self.json_to_csv,
            "query_json":         self.query_json,
            "validate_json":      self.validate_json_schema,
            # SQL / SQLite
            "connect_db":         self.connect_database,
            "disconnect_db":      self.disconnect_database,
            "execute_sql":        self.execute_sql,
            "query_sql":          self.query_sql,
            "list_tables":        self.list_db_tables,
            "describe_table":     self.describe_table,
            "import_csv_to_db":   self.import_csv_to_db,
            "export_table_to_csv": self.export_table_to_csv,
            "create_table":       self.create_table,
            "insert_rows":        self.insert_rows,
            "update_rows":        self.update_rows,
            "delete_rows":        self.delete_rows,
            "backup_db":          self.backup_database,
            # Analysis
            "profile":            self.profile_data,
            "describe":           self.describe_data,
            "correlate":          self.correlate_columns,
            "group_by":           self.group_by,
            "pivot":              self.pivot_table,
            "frequency":          self.frequency_table,
            "outliers":           self.detect_outliers,
            "missing":            self.analyze_missing,
            "normalize":          self.normalize_column,
            "encode_categorical": self.encode_categorical,
            "bin_column":         self.bin_column,
            "rolling_average":    self.rolling_average,
            # Transform
            "rename_columns":     self.rename_columns,
            "add_column":         self.add_column,
            "drop_columns":       self.drop_columns,
            "select_columns":     self.select_columns,
            "head":               self.head,
            "tail":               self.tail,
            "sample_data":        self.sample_rows,
            "cast_column":        self.cast_column,
            "fill_missing":       self.fill_missing,
            "drop_nulls":         self.drop_null_rows,
            "replace_values":     self.replace_values,
            "split_column":       self.split_column,
            "concat_columns":     self.concat_columns,
            "apply_formula":      self.apply_formula,
            # Charts
            "bar_chart":          self.bar_chart,
            "line_chart":         self.line_chart,
            "scatter_chart":      self.scatter_chart,
            "histogram":          self.histogram,
            "pie_chart":          self.pie_chart,
            "heatmap":            self.heatmap,
            "box_plot":           self.box_plot,
            # Misc
            "infer_schema":       self.infer_schema,
            "generate_dummy_data": self.generate_dummy_data,
            "compare_datasets":   self.compare_datasets,
            "get_query_log":      self._get_query_log,
        }


    # ─────────────────────────────────────────────────────────────────────────
    #  CSV Operations
    # ─────────────────────────────────────────────────────────────────────────

    def read_csv(self, path: str, delimiter: str = ",",
                  encoding: str = "utf-8", limit: int = None,
                  skip_rows: int = 0) -> Dict:
        """Read CSV file into list of dicts."""
        rows = self._read_csv_raw(path, delimiter, encoding, skip_rows)
        total = len(rows)
        if limit:
            rows = rows[:limit]
        return {
            "success":   True,
            "path":      path,
            "rows":      rows,
            "total":     total,
            "columns":   list(rows[0].keys()) if rows else [],
            "returned":  len(rows),
        }

    def write_csv(self, path: str, rows: List[Dict],
                   delimiter: str = ",",
                   encoding: str = "utf-8") -> Dict:
        if not rows:
            return {"success": False, "error": "No data to write"}
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", newline="", encoding=encoding) as f:
            w = csv.DictWriter(f, fieldnames=rows[0].keys(), delimiter=delimiter)
            w.writeheader()
            w.writerows(rows)
        return {"success": True, "path": str(path), "rows_written": len(rows)}

    def append_csv(self, path: str, rows: List[Dict],
                    delimiter: str = ",") -> Dict:
        p     = Path(path)
        exists = p.exists()
        with open(p, "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=rows[0].keys(), delimiter=delimiter)
            if not exists:
                w.writeheader()
            w.writerows(rows)
        return {"success": True, "path": str(p), "appended": len(rows)}

    def merge_csv(self, paths: List[str], output_path: str,
                   deduplicate: bool = False) -> Dict:
        """Merge multiple CSV files into one."""
        all_rows: List[Dict] = []
        for path in paths:
            rows = self._read_csv_raw(path)
            all_rows.extend(rows)

        if deduplicate:
            seen: set = set()
            deduped   = []
            for row in all_rows:
                key = json.dumps(row, sort_keys=True)
                if key not in seen:
                    seen.add(key); deduped.append(row)
            all_rows = deduped

        r = self.write_csv(output_path, all_rows)
        r["source_files"] = paths
        r["total_rows"]   = len(all_rows)
        return r

    def filter_csv(self, path: str, column: str, operator: str,
                    value: Any, output_path: str = None) -> Dict:
        """Filter rows where column <op> value. Ops: ==,!=,<,>,<=,>=,contains,startswith,endswith."""
        rows    = self._read_csv_raw(path)
        filtered = [r for r in rows if self._compare(r.get(column, ""), operator, value)]

        if output_path:
            self.write_csv(output_path, filtered)

        return {
            "success":    True,
            "filter":     f"{column} {operator} {value}",
            "in_rows":    len(rows),
            "out_rows":   len(filtered),
            "rows":       filtered[:1000],
            "saved_to":   output_path,
        }

    def sort_csv(self, path: str, column: str,
                  ascending: bool = True, output_path: str = None) -> Dict:
        rows = self._read_csv_raw(path)
        try:
            rows.sort(key=lambda r: self._coerce(r.get(column, "")),
                      reverse=not ascending)
        except Exception:
            rows.sort(key=lambda r: str(r.get(column, "")), reverse=not ascending)

        if output_path:
            self.write_csv(output_path, rows)
        return {"success": True, "sorted_by": column, "ascending": ascending,
                "rows": rows[:500], "total": len(rows)}

    def deduplicate_csv(self, path: str, subset: List[str] = None,
                         output_path: str = None) -> Dict:
        rows = self._read_csv_raw(path)
        seen: set = set()
        deduped   = []
        for row in rows:
            key = json.dumps({k: row[k] for k in (subset or row.keys())}, sort_keys=True)
            if key not in seen:
                seen.add(key); deduped.append(row)

        removed = len(rows) - len(deduped)
        if output_path:
            self.write_csv(output_path, deduped)
        return {"success": True, "original": len(rows), "deduped": len(deduped),
                "removed": removed, "saved_to": output_path}

    def validate_csv(self, path: str, required_columns: List[str] = None,
                      max_rows: int = None, check_types: Dict[str, str] = None) -> Dict:
        """Validate CSV structure and optionally type-check columns."""
        rows = self._read_csv_raw(path)
        if not rows:
            return {"success": False, "error": "Empty CSV", "valid": False}

        columns = list(rows[0].keys())
        errors: List[str] = []

        if required_columns:
            missing = [c for c in required_columns if c not in columns]
            if missing:
                errors.append(f"Missing columns: {missing}")

        if max_rows and len(rows) > max_rows:
            errors.append(f"Row count {len(rows)} exceeds limit {max_rows}")

        if check_types:
            for col, expected_type in check_types.items():
                if col not in columns:
                    continue
                for i, row in enumerate(rows, 1):
                    val = row.get(col, "")
                    try:
                        _TYPE_CHECKERS = {
                            "int":   lambda v: int(v),
                            "float": lambda v: float(v),
                            "date":  lambda v: datetime.fromisoformat(v),
                        }
                        checker = _TYPE_CHECKERS.get(expected_type)
                        checker(val) if checker else None
                    except (ValueError, TypeError):
                        errors.append(f"Row {i}, col '{col}': '{val}' not {expected_type}")
                        if len(errors) > 20:
                            errors.append("... (more errors truncated)")
                            break

        return {
            "success": True,
            "valid":   len(errors) == 0,
            "errors":  errors,
            "rows":    len(rows),
            "columns": columns,
        }

    def csv_to_json(self, csv_path: str, json_path: str = None,
                     indent: int = 2) -> Dict:
        rows = self._read_csv_raw(csv_path)
        js   = json.dumps(rows, indent=indent, ensure_ascii=False, default=str)
        if json_path:
            Path(json_path).write_text(js, encoding="utf-8")
        return {"success": True, "rows": len(rows),
                "json": js[:10000], "saved_to": json_path}

    # ─────────────────────────────────────────────────────────────────────────
    #  Excel Operations
    # ─────────────────────────────────────────────────────────────────────────

    def read_excel(self, path: str, sheet: str = None,
                    limit: int = None) -> Dict:
        if self._pd:
            try:
                df = self._pd.read_excel(path, sheet_name=sheet or 0,
                                           nrows=limit)
                rows = df.fillna("").to_dict("records")
                return {
                    "success": True,
                    "path":    path,
                    "sheet":   sheet or "Sheet1",
                    "rows":    rows[:5000],
                    "total":   len(rows),
                    "columns": list(df.columns),
                }
            except Exception as e:
                logger.warning(f"pandas read_excel failed: {e}")

        if self._xl:
            try:
                wb = self._xl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[sheet] if sheet else wb.active
                rows_raw = list(ws.iter_rows(values_only=True))
                if not rows_raw:
                    return {"success": True, "rows": [], "columns": []}
                headers = [str(c) if c is not None else f"col_{i}"
                           for i, c in enumerate(rows_raw[0])]
                rows    = [dict(zip(headers, row)) for row in rows_raw[1:]]
                if limit:
                    rows = rows[:limit]
                return {"success": True, "path": path, "sheet": ws.title,
                        "rows": rows, "total": len(rows), "columns": headers}
            except Exception as e:
                return {"success": False, "error": f"openpyxl error: {e}"}

        return {"success": False, "error": "pandas or openpyxl required for Excel"}

    def write_excel(self, path: str, data: Dict[str, List[Dict]],
                     auto_filter: bool = True,
                     freeze_header: bool = True) -> Dict:
        """Write multiple sheets. data = {sheet_name: [rows]}."""
        if not self._xl:
            return {"success": False, "error": "openpyxl required"}
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, rows in data.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            if not rows:
                continue
            headers = list(rows[0].keys())
            # Header row styling
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font        = Font(bold=True, color="FFFFFF")
                cell.fill        = PatternFill("solid", fgColor="4472C4")
                cell.alignment   = Alignment(horizontal="center")

            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            if auto_filter:
                ws.auto_filter.ref = ws.dimensions
            if freeze_header:
                ws.freeze_panes = "A2"

        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return {
            "success": True,
            "path":    path,
            "sheets":  list(data.keys()),
            "total_rows": sum(len(v) for v in data.values()),
        }

    def list_excel_sheets(self, path: str) -> Dict:
        if self._xl:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            return {"success": True, "sheets": wb.sheetnames}
        if self._pd:
            xl = self._pd.ExcelFile(path)
            return {"success": True, "sheets": xl.sheet_names}
        return {"success": False, "error": "openpyxl or pandas required"}

    def add_excel_sheet(self, path: str, sheet_name: str,
                         rows: List[Dict]) -> Dict:
        if not self._xl:
            return {"success": False, "error": "openpyxl required"}
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.create_sheet(title=sheet_name)
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(h, "") for h in headers])
        wb.save(path)
        return {"success": True, "path": path, "sheet": sheet_name, "rows": len(rows)}

    def read_excel_sheet(self, path: str, sheet: str,
                          limit: int = None) -> Dict:
        return self.read_excel(path, sheet=sheet, limit=limit)

    def excel_to_csv(self, excel_path: str, csv_path: str = None,
                      sheet: str = None) -> Dict:
        r = self.read_excel(excel_path, sheet=sheet)
        if not r["success"]:
            return r
        csv_path = csv_path or excel_path.rsplit(".", 1)[0] + ".csv"
        return self.write_csv(csv_path, r["rows"])

    def csv_to_excel(self, csv_path: str, excel_path: str = None,
                      sheet_name: str = "Sheet1") -> Dict:
        r = self.read_csv(csv_path)
        if not r["success"]:
            return r
        excel_path = excel_path or csv_path.rsplit(".", 1)[0] + ".xlsx"
        return self.write_excel(excel_path, {sheet_name: r["rows"]})

    # ─────────────────────────────────────────────────────────────────────────
    #  JSON Operations
    # ─────────────────────────────────────────────────────────────────────────

    def read_json(self, path: str, jq_path: str = None) -> Dict:
        """Read JSON file. Optionally apply a dot-path to extract nested data."""
        text = Path(path).read_text(encoding="utf-8")
        data = json.loads(text)
        if jq_path:
            data = self._traverse_json(data, jq_path)
        return {"success": True, "path": path, "data": data,
                "type": type(data).__name__}

    def write_json(self, path: str, data: Any, indent: int = 2,
                    ensure_ascii: bool = False) -> Dict:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        text = json.dumps(data, indent=indent, ensure_ascii=ensure_ascii,
                          default=str)
        Path(path).write_text(text, encoding="utf-8")
        return {"success": True, "path": path, "bytes": len(text)}

    def flatten_json(self, data: Union[Dict, str],
                      separator: str = ".") -> Dict:
        """Flatten nested JSON dict to single-level."""
        if isinstance(data, str):
            data = json.loads(data)
        flat = self._flatten_dict(data, sep=separator)
        return {"success": True, "flat": flat, "keys": len(flat)}

    def json_to_csv(self, json_path: str = None, csv_path: str = None,
                     data: List[Dict] = None) -> Dict:
        if json_path:
            r    = self.read_json(json_path)
            data = r["data"]
        if not isinstance(data, list):
            data = [data] if isinstance(data, dict) else []
        csv_path = csv_path or (json_path.rsplit(".", 1)[0] + ".csv" if json_path else None)
        if csv_path and data:
            self.write_csv(csv_path, data)
        return {"success": True, "rows": len(data or []), "saved_to": csv_path}

    def query_json(self, data: Union[List, Dict, str],
                    where: str = None,
                    select: List[str] = None,
                    order_by: str = None,
                    limit: int = None) -> Dict:
        """Simple SQL-like queries on JSON arrays."""
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except Exception:
                p = Path(data)
                data = json.loads(p.read_text()) if p.exists() else []

        if not isinstance(data, list):
            data = [data]

        # WHERE
        if where:
            data = self._apply_where_filter(data, where)

        # SELECT
        if select:
            data = [{k: r.get(k) for k in select} for r in data]

        # ORDER BY
        if order_by:
            reverse = order_by.startswith("-")
            col     = order_by.lstrip("-")
            data.sort(key=lambda r: r.get(col) or "", reverse=reverse)

        # LIMIT
        total = len(data)
        if limit:
            data = data[:limit]

        return {"success": True, "total": total, "returned": len(data), "data": data}

    def validate_json_schema(self, data: Any, schema: Dict) -> Dict:
        """Validate data against a simple schema dict of {field: type_str}."""
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except Exception:
                return {"success": False, "error": "Invalid JSON string"}

        records = data if isinstance(data, list) else [data]
        errors: List[str] = []
        type_map = {"str": str, "int": int, "float": float,
                    "bool": bool, "list": list, "dict": dict}

        for i, rec in enumerate(records):
            for field_name, type_str in schema.items():
                val = rec.get(field_name)
                if val is None:
                    errors.append(f"Row {i}: missing field '{field_name}'")
                elif type_str in type_map:
                    if not isinstance(val, type_map[type_str]):
                        errors.append(f"Row {i}: '{field_name}' expected {type_str}, got {type(val).__name__}")

        return {
            "success": True,
            "valid":   len(errors) == 0,
            "errors":  errors[:50],
            "checked": len(records),
        }

    # ─────────────────────────────────────────────────────────────────────────
    #  SQL / SQLite
    # ─────────────────────────────────────────────────────────────────────────

    def connect_database(self, path: str = ":memory:",
                          dsn: str = None) -> Dict:
        """Connect to SQLite (default) or any SQLAlchemy DSN."""
        if dsn and self._sa:
            from sqlalchemy import create_engine, text
            self._sa_engine = create_engine(dsn)
            self._db_path   = dsn
            return {"success": True, "type": "sqlalchemy", "dsn": dsn}

        # SQLite
        if path == ":memory:" or not Path(path).parent.exists():
            if path != ":memory:":
                Path(path).parent.mkdir(parents=True, exist_ok=True)
        self._active_db = sqlite3.connect(path, check_same_thread=False)
        self._active_db.row_factory = sqlite3.Row
        self._db_path   = path
        return {"success": True, "type": "sqlite", "path": path}

    def disconnect_database(self) -> Dict:
        if self._active_db:
            self._active_db.close()
            self._active_db = None
        return {"success": True, "disconnected": self._db_path}

    def execute_sql(self, sql: str, params: List = None,
                     commit: bool = True) -> Dict:
        """Execute a non-SELECT statement (INSERT/UPDATE/DELETE/CREATE)."""
        self._ensure_db()
        start = datetime.now()
        try:
            cur = self._active_db.cursor()
            cur.execute(sql, params or [])
            if commit:
                self._active_db.commit()
            elapsed = (datetime.now() - start).total_seconds()
            self._log_query(sql, elapsed)
            return {
                "success":      True,
                "rows_affected": cur.rowcount,
                "last_row_id":  cur.lastrowid,
                "elapsed_ms":   round(elapsed * 1000, 2),
            }
        except sqlite3.Error as e:
            return {"success": False, "error": str(e), "sql": sql[:200]}

    def query_sql(self, sql: str, params: List = None,
                   limit: int = 1000) -> Dict:
        """Execute a SELECT and return rows as list of dicts."""
        self._ensure_db()
        start = datetime.now()
        try:
            cur = self._active_db.cursor()
            cur.execute(sql, params or [])
            cols  = [d[0] for d in cur.description]
            rows  = [dict(zip(cols, r)) for r in cur.fetchmany(limit)]
            elapsed = (datetime.now() - start).total_seconds()
            self._log_query(sql, elapsed)
            return {
                "success":   True,
                "columns":   cols,
                "rows":      rows,
                "count":     len(rows),
                "elapsed_ms": round(elapsed * 1000, 2),
            }
        except sqlite3.Error as e:
            return {"success": False, "error": str(e), "sql": sql[:200]}

    def list_db_tables(self) -> Dict:
        self._ensure_db()
        r = self.query_sql("SELECT name, type FROM sqlite_master WHERE type IN ('table','view') ORDER BY name")
        return {
            "success": r["success"],
            "tables":  [row["name"] for row in r.get("rows", []) if row["type"] == "table"],
            "views":   [row["name"] for row in r.get("rows", []) if row["type"] == "view"],
        }

    def describe_table(self, table: str) -> Dict:
        self._ensure_db()
        schema_r = self.query_sql(f"PRAGMA table_info('{table}')")
        idx_r    = self.query_sql(f"PRAGMA index_list('{table}')")
        count_r  = self.query_sql(f"SELECT COUNT(*) AS cnt FROM \"{table}\"")
        return {
            "success": True,
            "table":   table,
            "columns": schema_r.get("rows", []),
            "indexes": idx_r.get("rows", []),
            "row_count": count_r["rows"][0]["cnt"] if count_r["success"] else None,
        }

    def import_csv_to_db(self, csv_path: str, table_name: str,
                          create_table: bool = True,
                          batch_size: int = 500) -> Dict:
        """Import a CSV file into a SQLite table."""
        self._ensure_db()
        rows = self._read_csv_raw(csv_path)
        if not rows:
            return {"success": False, "error": "CSV is empty"}

        columns = list(rows[0].keys())

        if create_table:
            # Infer types
            types   = self._infer_column_types(rows, columns)
            col_defs = ", ".join(f'"{c}" {t}' for c, t in zip(columns, types))
            self.execute_sql(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({col_defs})')

        placeholders = ", ".join("?" * len(columns))
        insert_sql   = f'INSERT INTO "{table_name}" VALUES ({placeholders})'
        cur = self._active_db.cursor()
        inserted = 0
        for i in range(0, len(rows), batch_size):
            batch = [[row.get(c, None) for c in columns] for row in rows[i:i + batch_size]]
            cur.executemany(insert_sql, batch)
            self._active_db.commit()
            inserted += len(batch)

        return {
            "success":  True,
            "table":    table_name,
            "inserted": inserted,
            "columns":  columns,
        }

    def export_table_to_csv(self, table_name: str, csv_path: str,
                              where: str = None) -> Dict:
        sql = f'SELECT * FROM "{table_name}"'
        if where:
            sql += f" WHERE {where}"
        r = self.query_sql(sql, limit=1_000_000)
        if not r["success"]:
            return r
        self.write_csv(csv_path, r["rows"])
        return {"success": True, "table": table_name, "rows": len(r["rows"]),
                "saved_to": csv_path}

    def create_table(self, table_name: str,
                      columns: Dict[str, str],
                      primary_key: str = None,
                      if_not_exists: bool = True) -> Dict:
        col_parts = []
        for name, dtype in columns.items():
            part = f'"{name}" {dtype}'
            if name == primary_key:
                part += " PRIMARY KEY"
            col_parts.append(part)
        exists = "IF NOT EXISTS " if if_not_exists else ""
        sql    = f'CREATE TABLE {exists}"{table_name}" ({", ".join(col_parts)})'
        return self.execute_sql(sql)

    def insert_rows(self, table_name: str, rows: List[Dict]) -> Dict:
        if not rows:
            return {"success": False, "error": "No rows to insert"}
        self._ensure_db()
        cols     = list(rows[0].keys())
        ph       = ", ".join("?" * len(cols))
        col_list = ", ".join('"' + c + '"' for c in cols)
        sql      = f'INSERT INTO "{table_name}" ({col_list}) VALUES ({ph})'
        cur   = self._active_db.cursor()
        data  = [[r.get(c) for c in cols] for r in rows]
        cur.executemany(sql, data)
        self._active_db.commit()
        return {"success": True, "inserted": len(rows), "table": table_name}

    def update_rows(self, table_name: str, values: Dict,
                     where: str, params: List = None) -> Dict:
        set_clause = ", ".join(f'"{k}" = ?' for k in values)
        sql        = f'UPDATE "{table_name}" SET {set_clause} WHERE {where}'
        all_params = list(values.values()) + (params or [])
        return self.execute_sql(sql, all_params)

    def delete_rows(self, table_name: str, where: str,
                     params: List = None) -> Dict:
        sql = f'DELETE FROM "{table_name}" WHERE {where}'
        return self.execute_sql(sql, params)

    def backup_database(self, backup_path: str) -> Dict:
        if not self._active_db:
            return {"success": False, "error": "No database connected"}
        backup_conn = sqlite3.connect(backup_path)
        self._active_db.backup(backup_conn)
        backup_conn.close()
        size = Path(backup_path).stat().st_size
        return {"success": True, "backup_path": backup_path,
                "size_kb": round(size / 1024, 1)}

    # ─────────────────────────────────────────────────────────────────────────
    #  Analysis
    # ─────────────────────────────────────────────────────────────────────────

    def profile_data(self, rows: List[Dict] = None,
                      path: str = None) -> Dict:
        """Comprehensive data profile: stats, types, nulls, uniques."""
        if path:
            rows = self._load_any(path)
        if not rows:
            return {"success": False, "error": "No data"}

        columns = list(rows[0].keys())
        n       = len(rows)

        dtypes:       Dict = {}
        null_counts:  Dict = {}
        unique_counts: Dict = {}
        numeric_stats: Dict = {}

        for col in columns:
            vals       = [r.get(col) for r in rows]
            non_null   = [v for v in vals if v is not None and str(v).strip() != ""]
            null_counts[col]   = n - len(non_null)
            unique_counts[col] = len(set(str(v) for v in non_null))

            nums = []
            for v in non_null:
                try:
                    nums.append(float(v))
                except (ValueError, TypeError):
                    pass

            if len(nums) > len(non_null) * 0.5:
                dtypes[col] = "numeric"
                if nums:
                    numeric_stats[col] = {
                        "min":    min(nums),
                        "max":    max(nums),
                        "mean":   statistics.mean(nums),
                        "median": statistics.median(nums),
                        "stdev":  statistics.stdev(nums) if len(nums) > 1 else 0,
                        "sum":    sum(nums),
                        "count":  len(nums),
                    }
            else:
                dtypes[col] = "string"

        return {
            "success":       True,
            "rows":          n,
            "columns":       len(columns),
            "column_names":  columns,
            "dtypes":        dtypes,
            "null_counts":   null_counts,
            "unique_counts": unique_counts,
            "numeric_stats": numeric_stats,
            "sample":        rows[:5],
            "null_pct":      {c: round(null_counts[c] / n * 100, 1) for c in columns},
        }

    def describe_data(self, rows: List[Dict], columns: List[str] = None) -> Dict:
        """Return descriptive stats for numeric columns."""
        if columns:
            rows = [{k: v for k, v in r.items() if k in columns} for r in rows]
        profile = self.profile_data(rows)
        return {
            "success":  profile["success"],
            "stats":    profile.get("numeric_stats", {}),
            "dtypes":   profile.get("dtypes", {}),
            "row_count": profile.get("rows", 0),
        }

    def correlate_columns(self, rows: List[Dict],
                           col_a: str, col_b: str) -> Dict:
        """Pearson correlation between two numeric columns."""
        try:
            a_vals = [float(r[col_a]) for r in rows if r.get(col_a) not in (None, "")]
            b_vals = [float(r[col_b]) for r in rows if r.get(col_b) not in (None, "")]
            n = min(len(a_vals), len(b_vals))
            if n < 2:
                return {"success": False, "error": "Not enough numeric values"}
            a, b = a_vals[:n], b_vals[:n]
            mean_a, mean_b = statistics.mean(a), statistics.mean(b)
            cov   = sum((x - mean_a) * (y - mean_b) for x, y in zip(a, b)) / (n - 1)
            std_a = statistics.stdev(a) or 1e-9
            std_b = statistics.stdev(b) or 1e-9
            corr  = cov / (std_a * std_b)
            return {
                "success":     True,
                "col_a":       col_a,
                "col_b":       col_b,
                "pearson_r":   round(corr, 4),
                "n_pairs":     n,
                "interpretation": (
                    "strong positive" if corr > 0.7 else
                    "moderate positive" if corr > 0.4 else
                    "weak positive" if corr > 0.1 else
                    "no correlation" if abs(corr) <= 0.1 else
                    "weak negative" if corr > -0.4 else
                    "moderate negative" if corr > -0.7 else
                    "strong negative"
                ),
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def group_by(self, rows: List[Dict], group_col: str,
                  agg_col: str = None, agg_func: str = "count") -> Dict:
        """Group rows and aggregate."""
        groups: Dict = defaultdict(list)
        for row in rows:
            key = str(row.get(group_col, ""))
            groups[key].append(row)

        result: List[Dict] = []
        for key, g_rows in sorted(groups.items()):
            entry: Dict = {group_col: key, "count": len(g_rows)}
            if agg_col:
                vals = [self._to_float(r.get(agg_col)) for r in g_rows
                        if r.get(agg_col) is not None]
                if vals:
                    entry.update({
                        f"{agg_func}_{agg_col}": {
                            "sum":    sum(vals),
                            "mean":   statistics.mean(vals),
                            "min":    min(vals),
                            "max":    max(vals),
                            "count":  len(vals),
                        }[agg_func] if agg_func in ("sum", "mean", "min", "max", "count")
                        else sum(vals),
                    })
            result.append(entry)

        return {"success": True, "groups": result, "group_count": len(result)}

    def pivot_table(self, rows: List[Dict], index: str,
                     column: str, value: str,
                     agg_func: str = "sum") -> Dict:
        """Create a pivot table."""
        index_vals  = sorted(set(str(r.get(index, ""))  for r in rows))
        column_vals = sorted(set(str(r.get(column, "")) for r in rows))

        # Bucket rows
        buckets: Dict = defaultdict(list)
        for row in rows:
            k = (str(row.get(index, "")), str(row.get(column, "")))
            v = self._to_float(row.get(value))
            if v is not None:
                buckets[k].append(v)

        # Build pivot
        pivot: List[Dict] = []
        for iv in index_vals:
            row_d: Dict = {index: iv}
            for cv in column_vals:
                vals = buckets.get((iv, cv), [])
                _AGG_DISPATCH = {
                    "sum":   sum,
                    "mean":  statistics.mean,
                    "count": len,
                    "min":   min,
                    "max":   max,
                }
                agg_fn = _AGG_DISPATCH.get(agg_func)
                row_d[cv] = agg_fn(vals) if agg_fn else None
            pivot.append(row_d)

        return {
            "success":      True,
            "pivot":        pivot,
            "index_values": index_vals,
            "columns":      column_vals,
        }

    def frequency_table(self, rows: List[Dict], column: str,
                          normalize: bool = False) -> Dict:
        """Value frequency count for a column."""
        vals = [str(r.get(column, "")) for r in rows]
        counter = Counter(vals)
        total   = len(vals)
        table   = [
            {
                "value":    v,
                "count":    c,
                "percent":  round(c / total * 100, 2) if normalize else None,
            }
            for v, c in counter.most_common()
        ]
        return {"success": True, "column": column, "total": total,
                "unique": len(counter), "table": table}

    def detect_outliers(self, rows: List[Dict], column: str,
                          method: str = "iqr",
                          threshold: float = 1.5) -> Dict:
        """Detect outliers using IQR or Z-score method."""
        vals = sorted([self._to_float(r.get(column)) for r in rows
                       if self._to_float(r.get(column)) is not None])
        if len(vals) < 4:
            return {"success": False, "error": "Not enough numeric values"}

        outlier_rows: List[Dict] = []

        def _iqr():
            q1, q3 = vals[len(vals)//4], vals[3*len(vals)//4]
            iqr = q3 - q1
            low, high = q1 - threshold * iqr, q3 + threshold * iqr
            for row in rows:
                v = self._to_float(row.get(column))
                if v is not None and (v < low or v > high): outlier_rows.append(row)

        def _zscore():
            mean_v, std_v = statistics.mean(vals), statistics.stdev(vals) or 1e-9
            for row in rows:
                v = self._to_float(row.get(column))
                if v is not None and abs((v - mean_v) / std_v) > threshold: outlier_rows.append(row)

        _OUTLIER_DISPATCH = {"iqr": _iqr, "zscore": _zscore}
        handler = _OUTLIER_DISPATCH.get(method.lower())
        handler() if handler else None

        return {
            "success":        True,
            "method":         method,
            "column":         column,
            "outlier_count":  len(outlier_rows),
            "total_rows":     len(rows),
            "outliers":       outlier_rows[:100],
        }

    def analyze_missing(self, rows: List[Dict]) -> Dict:
        """Analyse missing values across all columns."""
        if not rows:
            return {"success": False, "error": "No data"}
        columns = list(rows[0].keys())
        n = len(rows)
        missing: Dict = {}
        for col in columns:
            null_n = sum(1 for r in rows if r.get(col) in (None, "", "null", "NULL", "NaN"))
            missing[col] = {
                "count":   null_n,
                "percent": round(null_n / n * 100, 1),
            }
        return {"success": True, "rows": n, "missing": missing}

    def normalize_column(self, rows: List[Dict], column: str,
                           method: str = "minmax") -> Dict:
        """Normalize numeric column (minmax 0-1 or zscore)."""
        vals = [self._to_float(r.get(column)) for r in rows]
        nums = [v for v in vals if v is not None]
        if not nums:
            return {"success": False, "error": "No numeric values"}

        if method == "minmax":
            mn, mx = min(nums), max(nums)
            rng    = mx - mn or 1e-9
            normalized = [(v - mn) / rng if v is not None else None for v in vals]
        else:  # zscore
            mean_v = statistics.mean(nums)
            std_v  = statistics.stdev(nums) or 1e-9
            normalized = [(v - mean_v) / std_v if v is not None else None for v in vals]

        new_col = f"{column}_{method}"
        for row, nv in zip(rows, normalized):
            row[new_col] = round(nv, 6) if nv is not None else None

        return {"success": True, "column": column, "new_column": new_col,
                "method": method, "rows": rows[:1000]}

    def encode_categorical(self, rows: List[Dict], column: str,
                            method: str = "label") -> Dict:
        """Encode categorical column: 'label' (integer) or 'onehot'."""
        vals   = [str(r.get(column, "")) for r in rows]
        unique = sorted(set(vals))

        def _label():
            mapping = {v: i for i, v in enumerate(unique)}
            new_col = f"{column}_encoded"
            for row in rows: row[new_col] = mapping.get(str(row.get(column, "")), -1)
            return {"success": True, "mapping": mapping, "new_column": new_col, "rows": rows[:1000]}

        def _onehot():
            for cat in unique:
                col_name = f"{column}_{cat}"
                for row in rows: row[col_name] = 1 if str(row.get(column, "")) == cat else 0
            return {"success": True, "new_columns": [f"{column}_{c}" for c in unique], "categories": unique, "rows": rows[:1000]}

        _ENCODE_DISPATCH = {"label": _label, "onehot": _onehot}
        handler = _ENCODE_DISPATCH.get(method.lower())
        if not handler: return {"success": False, "error": f"Unknown method: {method}"}
        return handler()

    def bin_column(self, rows: List[Dict], column: str,
                    bins: int = 5, labels: List[str] = None) -> Dict:
        """Bin a numeric column into equal-width intervals."""
        vals = [self._to_float(r.get(column)) for r in rows]
        nums = [v for v in vals if v is not None]
        if not nums:
            return {"success": False, "error": "No numeric values"}

        mn, mx  = min(nums), max(nums)
        width   = (mx - mn) / bins or 1
        boundaries = [mn + i * width for i in range(bins + 1)]
        labels     = labels or [f"bin_{i+1}" for i in range(bins)]

        def _bin(v):
            if v is None:
                return None
            for i in range(bins):
                if v <= boundaries[i + 1]:
                    return labels[i]
            return labels[-1]

        new_col = f"{column}_bin"
        for row, v in zip(rows, vals):
            row[new_col] = _bin(v)

        return {"success": True, "column": column, "new_column": new_col,
                "bins": bins, "rows": rows[:1000]}

    def rolling_average(self, rows: List[Dict], column: str,
                          window: int = 3,
                          min_periods: int = 1) -> Dict:
        """Compute rolling (moving) average."""
        vals    = [self._to_float(r.get(column)) for r in rows]
        result  = []
        new_col = f"{column}_rolling_{window}"
        for i, (row, v) in enumerate(zip(rows, vals)):
            win    = [x for x in vals[max(0, i - window + 1): i + 1] if x is not None]
            if len(win) >= min_periods:
                row[new_col] = round(statistics.mean(win), 4)
            else:
                row[new_col] = None
        return {"success": True, "window": window, "new_column": new_col,
                "rows": rows[:1000]}

    # ─────────────────────────────────────────────────────────────────────────
    #  Transform
    # ─────────────────────────────────────────────────────────────────────────

    def rename_columns(self, rows: List[Dict],
                        mapping: Dict[str, str]) -> Dict:
        renamed = [{mapping.get(k, k): v for k, v in r.items()} for r in rows]
        return {"success": True, "mapping": mapping, "rows": renamed}

    def add_column(self, rows: List[Dict], column: str,
                    value: Any = None, formula: str = None) -> Dict:
        """Add a constant or formula-based column. Formula uses safe AST evaluator."""
        for row in rows:
            if formula:
                row[column] = _safe_eval_formula(formula, row)
            else:
                row[column] = value
        return {"success": True, "column": column, "rows": rows[:1000]}

    def drop_columns(self, rows: List[Dict], columns: List[str]) -> Dict:
        cleaned = [{k: v for k, v in r.items() if k not in columns} for r in rows]
        return {"success": True, "dropped": columns, "rows": cleaned}

    def select_columns(self, rows: List[Dict], columns: List[str]) -> Dict:
        selected = [{k: r.get(k) for k in columns} for r in rows]
        return {"success": True, "columns": columns, "rows": selected}

    def head(self, rows: List[Dict], n: int = 10) -> Dict:
        return {"success": True, "rows": rows[:n], "total": len(rows)}

    def tail(self, rows: List[Dict], n: int = 10) -> Dict:
        return {"success": True, "rows": rows[-n:], "total": len(rows)}

    def sample_rows(self, rows: List[Dict], n: int = 10,
                     seed: int = 42) -> Dict:
        import random
        rng    = random.Random(seed)
        sample = rng.sample(rows, min(n, len(rows)))
        return {"success": True, "rows": sample, "n": len(sample)}

    def cast_column(self, rows: List[Dict], column: str,
                     to_type: str) -> Dict:
        """Cast column to int/float/str/bool."""
        type_fn = {"int": int, "float": float, "str": str,
                    "bool": lambda x: x.lower() in ("1", "true", "yes")}
        fn = type_fn.get(to_type.lower())
        if not fn:
            return {"success": False, "error": f"Unknown type: {to_type}"}
        errors = 0
        for row in rows:
            try:
                row[column] = fn(row[column]) if row.get(column) is not None else None
            except (ValueError, TypeError, AttributeError):
                row[column] = None
                errors += 1
        return {"success": True, "column": column, "type": to_type,
                "cast_errors": errors, "rows": rows[:1000]}

    def fill_missing(self, rows: List[Dict], column: str,
                      strategy: str = "constant", value: Any = None) -> Dict:
        """Fill missing values: constant / mean / median / mode / forward / backward."""
        vals = [r.get(column) for r in rows]
        _STRAT_DISPATCH = {
            "mean":   lambda: statistics.mean([self._to_float(v) for v in vals if v not in (None, "")]) if any(v not in (None, "") for v in vals) else 0,
            "median": lambda: statistics.median(sorted(self._to_float(v) for v in vals if v not in (None, ""))) if any(v not in (None, "") for v in vals) else 0,
            "mode":   lambda: Counter([v for v in vals if v not in (None, "")]).most_common(1)[0][0] if any(v not in (None, "") for v in vals) else "",
        }
        
        handler = _STRAT_DISPATCH.get(strategy)
        fill = handler() if handler else value

        prev = fill
        for i, row in enumerate(rows):
            if row.get(column) in (None, ""):
                _FILL_MAP = {
                    "forward": lambda: prev,
                    "backward": lambda: fill,
                }
                row[column] = _FILL_MAP.get(strategy, lambda: fill)()
            else:
                prev = row[column]

        return {"success": True, "column": column, "strategy": strategy,
                "fill_value": str(fill), "rows": rows[:1000]}

    def drop_null_rows(self, rows: List[Dict],
                        columns: List[str] = None) -> Dict:
        """Drop rows with any null value in specified (or all) columns."""
        if columns:
            clean = [r for r in rows if all(r.get(c) not in (None, "") for c in columns)]
        else:
            clean = [r for r in rows if all(v not in (None, "") for v in r.values())]
        return {"success": True, "original": len(rows), "remaining": len(clean),
                "dropped": len(rows) - len(clean), "rows": clean}

    def replace_values(self, rows: List[Dict], column: str,
                        old_value: Any, new_value: Any) -> Dict:
        replaced = 0
        for row in rows:
            if str(row.get(column, "")) == str(old_value):
                row[column] = new_value
                replaced += 1
        return {"success": True, "column": column, "replaced": replaced,
                "rows": rows[:1000]}

    def split_column(self, rows: List[Dict], column: str,
                      delimiter: str = ",",
                      new_columns: List[str] = None) -> Dict:
        """Split a column's values by delimiter into multiple columns."""
        max_parts = 0
        for row in rows:
            parts = str(row.get(column, "")).split(delimiter)
            max_parts = max(max_parts, len(parts))

        cols = new_columns or [f"{column}_{i+1}" for i in range(max_parts)]
        for row in rows:
            parts = str(row.get(column, "")).split(delimiter)
            for i, c in enumerate(cols):
                row[c] = parts[i].strip() if i < len(parts) else ""

        return {"success": True, "column": column, "new_columns": cols,
                "rows": rows[:1000]}

    def concat_columns(self, rows: List[Dict], columns: List[str],
                        new_column: str,
                        separator: str = " ") -> Dict:
        for row in rows:
            row[new_column] = separator.join(str(row.get(c, "")) for c in columns)
        return {"success": True, "new_column": new_column, "rows": rows[:1000]}

    def apply_formula(self, rows: List[Dict], formula: str,
                       new_column: str) -> Dict:
        """Apply a safe expression to each row. Use column names as variables."""
        errors = 0
        for row in rows:
            result = _safe_eval_formula(formula, row)
            row[new_column] = result
            if result is None:
                errors += 1
        return {"success": True, "new_column": new_column, "errors": errors,
                "rows": rows[:1000]}

    # ─────────────────────────────────────────────────────────────────────────
    #  Charts (matplotlib)
    # ─────────────────────────────────────────────────────────────────────────

    def _save_chart(self, output_path: str = None, title: str = "") -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib not installed"}
        plt = self._plt
        plt.title(title)
        plt.tight_layout()
        if output_path:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            plt.savefig(output_path, dpi=150)
            plt.close()
            return {"success": True, "saved_to": output_path}
        else:
            buf = io.BytesIO()
            plt.savefig(buf, format="png", dpi=100)
            plt.close()
            buf.seek(0)
            import base64
            b64 = base64.b64encode(buf.read()).decode()
            return {"success": True, "png_base64": b64[:1000] + "..."}

    def bar_chart(self, rows: List[Dict], x_col: str, y_col: str,
                   title: str = "", output_path: str = None,
                   color: str = "steelblue") -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib required"}
        plt = self._plt
        plt.figure(figsize=(10, 6))
        x = [str(r.get(x_col, "")) for r in rows[:50]]
        y = [self._to_float(r.get(y_col)) or 0 for r in rows[:50]]
        plt.bar(x, y, color=color)
        plt.xlabel(x_col); plt.ylabel(y_col)
        plt.xticks(rotation=45, ha="right")
        return self._save_chart(output_path, title or f"{y_col} by {x_col}")

    def line_chart(self, rows: List[Dict], x_col: str, y_cols: List[str],
                    title: str = "", output_path: str = None) -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib required"}
        plt = self._plt
        plt.figure(figsize=(12, 6))
        x = range(len(rows))
        for col in y_cols:
            y = [self._to_float(r.get(col)) for r in rows]
            plt.plot(x, y, label=col, marker="o", markersize=3)
        plt.xlabel(x_col); plt.legend()
        return self._save_chart(output_path, title)

    def scatter_chart(self, rows: List[Dict], x_col: str, y_col: str,
                       color_col: str = None, title: str = "",
                       output_path: str = None) -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib required"}
        plt = self._plt
        plt.figure(figsize=(10, 7))
        x = [self._to_float(r.get(x_col)) for r in rows]
        y = [self._to_float(r.get(y_col)) for r in rows]
        valid = [(xi, yi) for xi, yi in zip(x, y) if xi is not None and yi is not None]
        if not valid:
            return {"success": False, "error": "No valid numeric pairs"}
        xv, yv = zip(*valid)
        plt.scatter(xv, yv, alpha=0.7, s=20)
        plt.xlabel(x_col); plt.ylabel(y_col)
        return self._save_chart(output_path, title or f"{x_col} vs {y_col}")

    def histogram(self, rows: List[Dict], column: str,
                   bins: int = 20, title: str = "",
                   output_path: str = None) -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib required"}
        plt = self._plt
        vals = [self._to_float(r.get(column)) for r in rows if self._to_float(r.get(column)) is not None]
        if not vals:
            return {"success": False, "error": "No numeric values"}
        plt.figure(figsize=(10, 6))
        plt.hist(vals, bins=bins, edgecolor="black", color="coral")
        plt.xlabel(column); plt.ylabel("Frequency")
        return self._save_chart(output_path, title or f"Distribution of {column}")

    def pie_chart(self, rows: List[Dict], label_col: str,
                   value_col: str, title: str = "",
                   output_path: str = None) -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib required"}
        plt = self._plt
        plt.figure(figsize=(9, 9))
        labels = [str(r.get(label_col, "")) for r in rows[:20]]
        values = [self._to_float(r.get(value_col)) or 0 for r in rows[:20]]
        plt.pie(values, labels=labels, autopct="%1.1f%%", startangle=140)
        return self._save_chart(output_path, title)

    def heatmap(self, rows: List[Dict], x_col: str, y_col: str,
                 value_col: str, title: str = "",
                 output_path: str = None) -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib required"}
        pivot_r = self.pivot_table(rows, x_col, y_col, value_col)
        if not pivot_r["success"]:
            return pivot_r
        plt  = self._plt
        pivot = pivot_r["pivot"]
        cols  = pivot_r["columns"]
        plt.figure(figsize=(12, 8))
        data = [[self._to_float(r.get(c)) or 0 for c in cols] for r in pivot]
        im = plt.imshow(data, aspect="auto", cmap="YlOrRd")
        plt.colorbar(im)
        plt.xticks(range(len(cols)), cols, rotation=45)
        plt.yticks(range(len(pivot)), [r[x_col] for r in pivot])
        return self._save_chart(output_path, title or f"{value_col} heatmap")

    def box_plot(self, rows: List[Dict], columns: List[str],
                  title: str = "", output_path: str = None) -> Dict:
        if not self._plt:
            return {"success": False, "error": "matplotlib required"}
        plt = self._plt
        data = []
        for col in columns:
            vals = [self._to_float(r.get(col)) for r in rows
                    if self._to_float(r.get(col)) is not None]
            data.append(vals)
        plt.figure(figsize=(max(8, len(columns) * 2), 7))
        plt.boxplot(data, labels=columns, patch_artist=True)
        plt.xticks(rotation=45, ha="right")
        return self._save_chart(output_path, title or "Box Plot")

    # ─────────────────────────────────────────────────────────────────────────
    #  Schema / Generate / Compare
    # ─────────────────────────────────────────────────────────────────────────

    def infer_schema(self, rows: List[Dict]) -> Dict:
        """Infer column types from data sample."""
        if not rows:
            return {"success": False, "error": "No data"}
        columns = list(rows[0].keys())
        schema  = {}
        for col in columns:
            vals = [r.get(col) for r in rows if r.get(col) not in (None, "")]
            types = self._infer_types_from_values(vals)
            schema[col] = types
        return {"success": True, "schema": schema}

    def generate_dummy_data(self, schema: Dict[str, str],
                             n: int = 100, seed: int = 0) -> Dict:
        """Generate realistic dummy data from a schema {col: type}."""
        import random
        rng   = random.Random(seed)
        words = ["alpha", "beta", "gamma", "delta", "epsilon", "zeta",
                 "theta", "iota", "kappa", "lambda", "mu", "nu", "xi"]
        rows: List[Dict] = []
        for i in range(n):
            row: Dict = {}
            for col, dtype in schema.items():
                _GEN_HANDLERS = {
                    "int":     lambda: rng.randint(0, 10000),
                    "integer": lambda: rng.randint(0, 10000),
                    "float":   lambda: round(rng.uniform(0, 1000), 2),
                    "decimal": lambda: round(rng.uniform(0, 1000), 2),
                    "numeric": lambda: round(rng.uniform(0, 1000), 2),
                    "bool":    lambda: rng.choice([True, False]),
                    "boolean": lambda: rng.choice([True, False]),
                    "date":    lambda: datetime.fromordinal(datetime(2020, 1, 1).toordinal() + rng.randint(0, 1500)).strftime("%Y-%m-%d"),
                    "email":   lambda: f"user{i+1}@example.com",
                    "name":    lambda: rng.choice(words).capitalize() + " " + rng.choice(words).capitalize(),
                    "str":     lambda: rng.choice(words).capitalize() + " " + rng.choice(words).capitalize(),
                    "string":  lambda: rng.choice(words).capitalize() + " " + rng.choice(words).capitalize(),
                    "text":    lambda: rng.choice(words).capitalize() + " " + rng.choice(words).capitalize(),
                    "category":lambda: rng.choice(["A", "B", "C", "D"]),
                    "phone":   lambda: f"+1-{rng.randint(200,999)}-{rng.randint(100,999)}-{rng.randint(1000,9999)}",
                }
                row[col] = _GEN_HANDLERS.get(dtype, lambda: f"value_{i+1}")()
            rows.append(row)
        return {"success": True, "rows": rows, "count": n, "schema": schema}

    def compare_datasets(self, rows_a: List[Dict], rows_b: List[Dict],
                          key_column: str = None) -> Dict:
        """Compare two datasets and report differences."""
        cols_a = set(rows_a[0].keys()) if rows_a else set()
        cols_b = set(rows_b[0].keys()) if rows_b else set()

        missing_cols = cols_a - cols_b
        extra_cols   = cols_b - cols_a
        common_cols  = cols_a & cols_b

        result = {
            "success":        True,
            "rows_a":         len(rows_a),
            "rows_b":         len(rows_b),
            "columns_in_a_not_b": list(missing_cols),
            "columns_in_b_not_a": list(extra_cols),
            "common_columns":     list(common_cols),
            "row_count_diff":     len(rows_a) - len(rows_b),
        }

        if key_column and key_column in common_cols:
            keys_a = {str(r.get(key_column)): r for r in rows_a}
            keys_b = {str(r.get(key_column)): r for r in rows_b}
            result["only_in_a"] = [k for k in keys_a if k not in keys_b][:20]
            result["only_in_b"] = [k for k in keys_b if k not in keys_a][:20]
            # Row differences
            diffs: List[Dict] = []
            for k in set(keys_a) & set(keys_b):
                ra, rb = keys_a[k], keys_b[k]
                changed = {c: (ra.get(c), rb.get(c)) for c in common_cols
                           if str(ra.get(c)) != str(rb.get(c))}
                if changed:
                    diffs.append({"key": k, "changes": changed})
            result["changed_rows"] = diffs[:50]
            result["changed_count"] = len(diffs)

        return result

    # ─────────────────────────────────────────────────────────────────────────
    #  Helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _read_csv_raw(self, path: str, delimiter: str = ",",
                       encoding: str = "utf-8", skip_rows: int = 0) -> List[Dict]:
        with open(path, "r", newline="", encoding=encoding, errors="replace") as f:
            for _ in range(skip_rows):
                next(f, None)
            reader = csv.DictReader(f, delimiter=delimiter)
            return [dict(row) for row in reader]

    def _load_any(self, path: str) -> List[Dict]:
        ext = Path(path).suffix.lower()
        _LOAD_DISPATCH = {
            ".csv":  lambda: self._read_csv_raw(path, delimiter=","),
            ".tsv":  lambda: self._read_csv_raw(path, delimiter="\t"),
            ".json": lambda: [json.loads(Path(path).read_text())] if not isinstance(json.loads(Path(path).read_text()), list) else json.loads(Path(path).read_text()),
            ".xlsx": lambda: self.read_excel(path).get("rows", []),
            ".xls":  lambda: self.read_excel(path).get("rows", []),
        }
        handler = _LOAD_DISPATCH.get(ext)
        return handler() if handler else []

    def _ensure_db(self):
        if not self._active_db:
            self.connect_database(":memory:")

    def _log_query(self, sql: str, elapsed: float):
        self._query_log.append({
            "ts": datetime.now().isoformat(),
            "sql": sql[:300],
            "ms":  round(elapsed * 1000, 1),
        })
        if len(self._query_log) > 2000:
            self._query_log = self._query_log[-1000:]

    def _get_query_log(self, limit: int = 50) -> Dict:
        return {"success": True, "log": self._query_log[-limit:]}

    @staticmethod
    def _compare(val: Any, op: str, target: Any) -> bool:
        sv = str(val).lower()
        st = str(target).lower()
        try:
            fv, ft = float(val), float(target)
            if op == "<":  return fv < ft
            if op == ">":  return fv > ft
            if op == "<=": return fv <= ft
            if op == ">=": return fv >= ft
        except (ValueError, TypeError):
            pass
        if op == "==":         return sv == st
        if op == "!=":         return sv != st
        if op == "contains":   return st in sv
        if op == "startswith": return sv.startswith(st)
        if op == "endswith":   return sv.endswith(st)
        return False

    @staticmethod
    def _to_float(val: Any) -> Optional[float]:
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _coerce(val: Any) -> Any:
        try:
            return float(val)
        except (ValueError, TypeError):
            return str(val)

    def _apply_where_filter(self, rows: List[Dict], where: str) -> List[Dict]:
        """Evaluate a simple Python boolean expression per row using secure AST evaluation."""
        result = []
        for row in rows:
            try:
                if _safe_eval_formula(where, row):
                    result.append(row)
            except Exception:
                pass
        return result

    @staticmethod
    def _flatten_dict(d: Dict, parent_key: str = "",
                       sep: str = ".") -> Dict:
        items: Dict = {}
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(DataAgent._flatten_dict(v, new_key, sep))
            elif isinstance(v, list):
                for i, item in enumerate(v):
                    if isinstance(item, dict):
                        items.update(DataAgent._flatten_dict(item, f"{new_key}[{i}]", sep))
                    else:
                        items[f"{new_key}[{i}]"] = item
            else:
                items[new_key] = v
        return items

    @staticmethod
    def _traverse_json(data: Any, path: str) -> Any:
        def _dict_part(d, p): return d.get(p)
        def _list_part(d, p): return d[int(p)] if p.isdigit() else None
        _TRAVERSE_DISPATCH = {dict: _dict_part, list: _list_part}
        
        for part in path.split("."):
            h = _TRAVERSE_DISPATCH.get(type(data))
            if not h: return None
            data = h(data, part)
        return data

    @staticmethod
    def _infer_types_from_values(vals: List) -> str:
        if not vals:
            return "string"
        ints   = sum(1 for v in vals if re.match(r"^-?\d+$", str(v)))
        floats = sum(1 for v in vals if re.match(r"^-?\d+\.\d+$", str(v)))
        dates  = sum(1 for v in vals if re.match(r"^\d{4}-\d{2}-\d{2}", str(v)))
        n = len(vals)
        if ints / n > 0.8:    return "INTEGER"
        if floats / n > 0.8:  return "REAL"
        if dates / n > 0.8:   return "TEXT"  # DATE as TEXT in SQLite
        return "TEXT"

    @staticmethod
    def _infer_column_types(rows: List[Dict], columns: List[str]) -> List[str]:
        types = []
        for col in columns:
            vals = [str(r.get(col, "")) for r in rows if r.get(col) not in (None, "")]
            types.append(DataAgent._infer_types_from_values(vals))
        return types

    @staticmethod
    def _try_import(module_name: str):
        import importlib
        try:
            return importlib.import_module(module_name)
        except ImportError:
            return None
